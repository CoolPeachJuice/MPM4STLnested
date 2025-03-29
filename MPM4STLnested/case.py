import copy
import matplotlib.pyplot as plt
import numpy as np
import time
from shapely.geometry import Polygon, GeometryCollection
from shapely.affinity import translate
import xml.etree.ElementTree as ET
from openpyxl import Workbook

from myset import NDSet
from STLnested_tree import Phi

#  给两个列表取并集，考虑R以及\运算
def bing(region1, region2):
    # 用来给两个集合取并集，考虑全集R，R的存法是region = ['R'] 存了一个字符R
    if region1[0] == 'R':
        region = region1
    elif region2[0] == 'R':
        region = region2
    else:  # 这种情况是两个都不是R，正常取并集即可
        if region1[0] == 'not':
            if region1[1] == 'R': # 如果1是['not', 'R']
                region = region2
            else: # 1不是['not', 'R']，后边是多边体
                if region2[0] == 'not':  # 都是not
                    if region2[1] == 'R': # 2是['not', 'R']
                        region = region1
                    else: # 后边都是多边体
                        region = ['not', region1[1].intersection(region2[1])]
                else:  # 1是not，2不是
                    region = ['not', region1[1].difference(region2[0])]
        elif region2[0] == 'not':  # 这时候region1肯定不是not了
            if region2[1] == 'R':  # 2是['not', 'R']
                region = region1
            else: # 2不是['not', 'R']，后边是多边体
                region = ['not', region2[1].difference(region1[0])]
        else:  # 俩都不是not不是R，正常取并集
            region = [region1[0].union(region2[0])]
    return region

#  给两个集合取交集，考虑R以及\运算
def jiao(region1, region2):
    # 用来给两个集合取交集，考虑全集R，R的存法是region = ['R'] 存了一个字符R
    if region1[0] == 'R':
        region = region2
    elif region2[0] == 'R':
        region = region1
    else:  # 这种情况是两个都不是R，正常取交集即可
        if region1[0] == 'not':
            if region2[0] == 'not':
                if region1[1] == 'R' or region2[1] == 'R': # 1或2是['not', 'R']的情况
                    region = ['not', 'R']
                else:
                    region = ['not', region1[1].union(region2[1])]
            else:
                if region1[1] == 'R': # 1是['not', 'R']的情况
                    region = ['not', 'R']
                else:
                    region = [region2[0].difference(region1[1])]
        elif region2[0] == 'not':  # 这时候region1肯定不是not了
            if region2[1] == 'R':  # 2是['not', 'R']的情况
                region = ['not', 'R']
            else:
                region = [region1[0].difference(region2[1])]
        else:  # 俩都不是not不是R，正常取交集
            # print(region1)
            # print(region2)
            region = [region1[0].intersection(region2[0])]
    return region

def daochu(start, ws, father=None, id=1):
    # start是一个Feasible_Set(), father是一个ET的对象，作为下一级的父节点
    if father == None:
        root = ET.Element('root', attrib={'id': f'{id}'})
        ws.cell(row=id, column=1, value=f'{id}')
        ws.cell(row=id, column=2, value=f'{start.feasible_set}')
        id = id+1
    else:
        root = father
    for i in range(len(start.child_list)):
        child = ET.SubElement(root, f'k{start.k+1}', attrib={'id': f'{id}'}, text=f'{start.child_list[i].sat}')
        ws.cell(row=id, column=1, value=f'{id}')
        ws.cell(row=id, column=2, value=f'{start.child_list[i].feasible_set}')
        id = id + 1
        if start.child_list[i].child_list != []:
            id = daochu(start.child_list[i], ws, child, id)  # 这个child作为下一级的父节点
    if father == None:  # 到最外边一层才导出，下边的不用导出
        tree = ET.ElementTree(root)
        # print("111111111111111111111111111111")
        tree.write('leaf_satisfaction_tree.xml')
    return id


class Feasible_Set():
    # 用来存储X_k^I
    def __init__(self, k, probability_phi, X, child_list=None):
        self.k = k
        self.I = probability_phi
        self.X = X
        self.child_list = child_list  # 我所有的后继的可行集
        self.leaf_satisfaction_list = []  # 我到达每一个后继所需要的leaf_satisfaction，和child_list里的后继一一对应
        if self.child_list != None:
            self.set_leaf_satisfaction_list()
        else:
            self.leaf_satisfaction_list = None

    # 重写print
    def __repr__(self):
        prt = f"{self.I}"
        return prt

    def set_leaf_satisfaction_list(self):
        for i in range(len(self.child_list)):
            leaf_satisfaction = []
            I_next = self.child_list[i].I
            for j in range(len(self.I)):
                if self.I[j] == I_next[j]:
                    leaf_satisfaction += [None]
                else:
                    leaf_satisfaction += [I_next[j][-1]]
            self.leaf_satisfaction_list += [leaf_satisfaction]

    # 随手写的简单的用来检查树有没有存好的函数
    def print_tree(self):
        if self.child_list == None:
            print(self.I)
            # print(self.k)
        else:
            print(self.I)
            # print(self.k)
            for i in range(len(self.child_list)):
                print(self.leaf_satisfaction_list[i])
                self.child_list[i].print_tree()
        return True


def one_step_set_backward(S, A, BU):  # 反向求一步集
    # S是一个n维的NDSet，A是一个n*n矩阵，BU是一个n维超矩形
    # 现在只能处理A=I的情况
    # print(S.sorted_points())
    # rS = S.transform_and_union_hyperrectangles(A, BU, auto_sparsify=False)
    rS = S.transform_and_union_hyperrectangles(A, BU, auto_sparsify=True)
    # print(rS.sorted_points())
    return [rS]

def feasible_set_tree(task, A, BU):
    sat_tree_task = task.all_p()
    return feasible_set_task_digui(sat_tree_task, task, A, BU)

def feasible_set_task_digui(root, task, A, BU):
    print(root.k)
    if root.child_list == []:  # 说明是已经完成的情况，那feasible_set就是R
        root.feasible_set = ['R']
        return root
    else:
        for i in range(len(root.child_list)):
            root.child_list[i] = feasible_set_task_digui(root.child_list[i], task, A, BU)
        # 现在root的所有的child的feasible_set都已经算好了，可以算root的feasible_set了
        # 算root的feasible_set
        linshi_list = []
        for i in range(len(root.child_list)):
            if root.child_list[i].feasible_set == ['R']:
                one_step_set = ['R']
            else:
                one_step_set = one_step_set_backward(root.child_list[i].feasible_set[0], A, BU)
            consistent_region = task.consistent_region(root.child_list[i])  # 注意，空集不一定是[]或['not', 'R']，应该是[NDSet]，但是这个NDSet是空的
            linshi_list += [jiao(consistent_region, one_step_set)]
        X_k_I_list = linshi_list[0]
        for list1 in linshi_list:
            X_k_I_list = bing(X_k_I_list, list1)
        root.feasible_set = X_k_I_list
        return root

def get_feasible_set_list(root, task, feasible_set_list=None):
    if feasible_set_list == None:
        feasible_set_list = []
        max_k = task.effective_horizon()[1]
        for i in range(max_k+2):
            feasible_set_list += [[]]
    if len(root.child_list) != 0:
        feasible_set_list[root.k] += [root.feasible_set]
        # print()
        # print(root.k)
        # print(root.sat)
        # print(root.feasible_set)
        for child in root.child_list:
            get_feasible_set_list(child, task, feasible_set_list)
    return feasible_set_list


if __name__ == "__main__":
    nd_set = NDSet(2, 0)
    nd_set.add_points_in_hyperrectangle([(6, 8), (6, 8)])
    o2 = Phi([0], [2], ['G'], [[nd_set]])
    nd_set2 = NDSet(2, 0)
    nd_set2.add_points_in_hyperrectangle([(3, 5), (3, 5)])
    task = Phi([0, 0], [6, 6], ['U', 'U'], [[nd_set2], o2], [['R'], ['R']])

    t1 = time.time()
    print(task)
    print(task.tree)
    print(task.effective_horizon())
    print("---------------------------------------------------------------")
    # print(task.all_p())
    print("---------------------------------------------------------------")
    max_k = task.effective_horizon()[1]
    A = np.array([[1, 0], [0, 1]])
    BU = [(-1,1), (-1,1)]
    root = feasible_set_tree(task, A, BU)
    print(time.time()-t1)

    # 创建一个新的工作簿
    wb = Workbook()
    # 获取活动的工作表
    ws = wb.active
    daochu(root, ws)
    print()
    # for row in ws.iter_rows(values_only=True):
    #     print(row)
    wb.save("feasible_sets.xlsx")

    # print(root)

    feasible_set_list = get_feasible_set_list(root, task)
    print(len(feasible_set_list))
    # for i in range(len(feasible_set_list)):
    #     set_i = ['not', 'R']
    #     for j in range(len(feasible_set_list[i])):
    #         set_i = bing(set_i, feasible_set_list[i][j])
    #     print(set_i)
    #     set_i[0].plot()